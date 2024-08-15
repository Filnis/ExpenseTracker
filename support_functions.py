from openpyxl import load_workbook
import pandas as pd
from decimal import *
import dash_bootstrap_components as dbc
from decimal import *
from dash import html
from datetime import datetime, timedelta, date

def get_table(workbook, worksheet, table_name):
    # read file
    wb = load_workbook(workbook, data_only=True)

    # access specific sheet
    ws = wb[worksheet]
    tab = {key : value for key, value in ws.tables.items()}
    #print(tab)

    mapping = {}

    for entry, data_boundary in ws.tables.items():
        #parse the data within the ref boundary
        data = ws[data_boundary]
        #extract the data 
        #the inner list comprehension gets the values for each cell in the table
        content = [[cell.value for cell in ent] 
                for ent in data
            ]
        
        header = content[0]
        
        #the contents ... excluding the header
        rest = content[1:]
        
        #create dataframe with the column names
        #and pair table name with dataframe
        df = pd.DataFrame(rest, columns = header)
        mapping[entry] = df
    
    wb.close()
    return mapping[table_name]

def float2dec(value):
    return Decimal(value).quantize(Decimal('0.01'))

#creates a Bootstrap card 
def create_card(header, body):
    card = dbc.Card([
        dbc.CardHeader(header),
        dbc.CardBody(html.H4(body, className="card-title"))
    ],  color = 'warning', className='text-center shadow my-3') 

    return card

def create_date_table2(start, end):
    df = pd.DataFrame({"Date": pd.date_range(start, end)})
    df["Day"] = df.Date.dt.day_name()
    df["Day_num"] = df.Date.dt.day
    #df["Week"] = df.Date.dt.weekofyear
    df["Quarter"] = df.Date.dt.quarter
    df["Month"] = df.Date.dt.month
    df["Year"] = df.Date.dt.year
    df["Year_half"] = (df.Quarter + 1) // 2
    return df

def createCardV2(value, legend, icon):
    card_icon = {
    "color": "white",
    "textAlign": "center",
    "fontSize": 30,
    "margin": "auto",
    }
    card = dbc.CardGroup(
        [
            dbc.Card(
                dbc.CardBody(
                    [
                        html.H5(value, className="card-title"),
                        html.P(legend, className="card-text",),
                    ]
                )
            ),
            dbc.Card(
                html.Div(className=icon, style=card_icon),
                className="bg-primary",
                style={"maxWidth": 75},
            ),
        ],
        className="mt-4",
    )
    return card